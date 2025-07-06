import openpyxl

ws = openpyxl.load_workbook('6789.xlsx')
# sheet1 = ws.get_sheet_names('Sheet1')
sheet1 = ws['Sheet1']
wb = ws.active
a1 = wb['B4']
print(sheet1)
print(a1.value)
for row in wb.iter_rows(values_only=True):  # values_only=True 表示只返回单元格的值
    print(row)
# 追加标题
wb.cell(row=1, column=wb.max_column + 1, value="是否高薪")
# 按行判断“分数”列，然后追加“等级”
for row in wb.iter_rows(min_row=2, max_row=wb.max_row):
    score = row[2].value  # 第3列是分数
    if score >= 500000:
        grade = "是"
    else:
        grade = "否"
    wb.cell(row=row[0].row, column=wb.max_column, value=grade)
for row in wb.iter_rows(values_only=True):  # values_only=True 表示只返回单元格的值
    print(row)
ws.save("test.xlsx")